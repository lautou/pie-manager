"""
Tests for GET /api/transactions/import/template — Excel template generation.

The single highest-value test here is the self-consistency round-trip: every example row the
template generates must independently pass validation (status="ok") when fed back through
/validate. This is a structural guard against the template ever drifting out of sync with the
validation rules in import_service.py.
"""
import io

import pytest
from openpyxl import load_workbook

from app.models import Broker, Portfolio, PortfolioAccount, Product
from app.services.import_service import TRANSACTION_COLUMNS


async def _seed_full_reference_data(db_session, uid_suffix):
    portfolio = Portfolio(name=f"TplPortfolio-{uid_suffix}")
    db_session.add(portfolio)
    await db_session.flush()
    broker = Broker(name=f"TplBroker-{uid_suffix}", currency="EUR")
    db_session.add(broker)
    await db_session.flush()
    db_session.add(PortfolioAccount(portfolio_id=portfolio.id, broker_id=broker.id))

    products = [
        Product(ticker=f"ETF.{uid_suffix}", name="ETF Monde", category="Actif", instrument_type="ETF", currency="EUR"),
        Product(ticker=f"USD.ACT.{uid_suffix}", name="US Stock", category="Actif", instrument_type="Action", currency="USD"),
        Product(ticker=f"LIQUIDITE.EURO.{uid_suffix}", name="Cash EUR", category="Actif", instrument_type="Cash", currency="EUR"),
        Product(ticker="JPYEUR=X", name="JPY", category="Actif", instrument_type="Cash", currency="EUR"),
        Product(ticker=f"OR.{uid_suffix}", name="Or physique", category="Actif", instrument_type="Or physique", currency="EUR"),
        Product(ticker=f"FRAIS.{uid_suffix}", name="Tenue de compte", category="Frais", fee_type="Tenue de compte", currency="EUR"),
    ]
    for p in products:
        db_session.add(p)
    await db_session.flush()
    return portfolio, broker


@pytest.mark.asyncio
async def test_download_template_returns_xlsx(client, db_session):
    suffix = id(db_session)
    await _seed_full_reference_data(db_session, suffix)

    r = await client.get("/api/transactions/import/template/modele_import_transactions.xlsx")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in r.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(r.content))
    assert "Transactions" in wb.sheetnames
    assert "Instructions" in wb.sheetnames
    header = [c.value for c in wb["Transactions"][1]]
    assert header == TRANSACTION_COLUMNS


@pytest.mark.asyncio
async def test_template_instructions_sheet_has_content(client, db_session):
    suffix = id(db_session)
    await _seed_full_reference_data(db_session, suffix)
    r = await client.get("/api/transactions/import/template/modele_import_transactions.xlsx")
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Instructions"]
    all_text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "Achat Or physique" in all_text
    assert "VALEUR TOTALE" in all_text


@pytest.mark.asyncio
async def test_template_example_rows_pass_self_consistency_validation(client, db_session):
    """The core guard: every generated example row must resolve to status="ok" when
    re-submitted through /validate — a structural check against template/validation drift."""
    suffix = id(db_session)
    await _seed_full_reference_data(db_session, suffix)

    r = await client.get("/api/transactions/import/template/modele_import_transactions.xlsx")
    assert r.status_code == 200

    files = {"file": ("template.xlsx", io.BytesIO(r.content),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    validate_r = await client.post("/api/transactions/import/validate", files=files)
    assert validate_r.status_code == 200
    body = validate_r.json()
    assert body["summary"]["total_rows"] > 0
    for row in body["rows"]:
        assert row["status"] == "ok", f"row {row['row_number']} ({row['sens']}): {row['errors']}"
    assert body["summary"]["errors"] == 0
    assert body["summary"]["duplicates"] == 0


@pytest.mark.asyncio
async def test_template_with_no_data_still_returns_valid_empty_workbook(client, db_session):
    """A fresh install with no portfolios/brokers/products yet must not crash the endpoint —
    it degrades to a template with headers only, no example rows."""
    r = await client.get("/api/transactions/import/template/modele_import_transactions.xlsx")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Transactions"]
    assert ws.max_row == 1  # header only


@pytest.mark.asyncio
async def test_template_with_account_but_no_products_produces_no_example_rows(client, db_session):
    """An account exists (so _pick_account succeeds) but zero products are seeded — every
    "if <product found>:" block in _build_example_rows must independently evaluate to False
    without raising, producing an otherwise-valid template with no example rows."""
    suffix = id(db_session)
    portfolio = Portfolio(name=f"NoProductsTpl-{suffix}")
    db_session.add(portfolio)
    await db_session.flush()
    broker = Broker(name=f"NoProductsBroker-{suffix}", currency="EUR")
    db_session.add(broker)
    await db_session.flush()
    db_session.add(PortfolioAccount(portfolio_id=portfolio.id, broker_id=broker.id))
    await db_session.flush()

    r = await client.get("/api/transactions/import/template/modele_import_transactions.xlsx")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Transactions"]
    assert ws.max_row == 1  # header only, no products to build an example row from


@pytest.mark.asyncio
async def test_template_with_partial_product_coverage_only_uses_whats_available(client, db_session):
    """An account with only a Frais product (no ETF/Action/Cash/Or physique) must still
    produce a valid, self-consistent template — every "product missing" branch in
    _build_example_rows is independently skipped rather than raising."""
    suffix = id(db_session)
    portfolio = Portfolio(name=f"PartialTpl-{suffix}")
    db_session.add(portfolio)
    await db_session.flush()
    broker = Broker(name=f"PartialBroker-{suffix}", currency="EUR")
    db_session.add(broker)
    await db_session.flush()
    db_session.add(PortfolioAccount(portfolio_id=portfolio.id, broker_id=broker.id))
    db_session.add(Product(ticker=f"FRAIS.ONLY.{suffix}", name="Frais isolé", category="Frais", currency="EUR"))
    await db_session.flush()

    r = await client.get("/api/transactions/import/template/modele_import_transactions.xlsx")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Transactions"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1  # only the Frais example row could be generated
    assert rows[0][2] == "Frais"
