"""
Tests for POST /api/transactions/import/validate — dry-run validation over HTTP,
including duplicate detection (in-DB and in-file) and malformed-file handling.
Per-Sens/per-rule validation logic itself is covered exhaustively in
test_import_service.py (pure unit tests, no HTTP) — this file focuses on what only
exists at the HTTP/parsing layer.
"""
import io
from datetime import date

import pytest
from openpyxl import Workbook

from app.models import Broker, Portfolio, PortfolioAccount, Product, Transaction
from app.services.import_service import TRANSACTION_COLUMNS


def _make_xlsx(rows: list[dict], sheet_name: str = "Transactions") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(TRANSACTION_COLUMNS)
    for row in rows:
        ws.append([row.get(col) for col in TRANSACTION_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_files(content: bytes):
    return {"file": ("import.xlsx", io.BytesIO(content),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}


async def _seed_basic(db_session, suffix):
    portfolio = Portfolio(name=f"ValPortfolio-{suffix}")
    db_session.add(portfolio)
    await db_session.flush()
    broker = Broker(name=f"ValBroker-{suffix}", currency="EUR")
    db_session.add(broker)
    await db_session.flush()
    db_session.add(PortfolioAccount(portfolio_id=portfolio.id, broker_id=broker.id))
    product = Product(ticker=f"ETF.{suffix}", name="ETF Monde", category="Actif", instrument_type="ETF", currency="EUR")
    db_session.add(product)
    await db_session.flush()
    return portfolio, broker, product


def _row(portfolio, broker, product, **overrides):
    base = {
        "Portefeuille": portfolio.name, "Compte": broker.name, "Sens": "Achat",
        "Ticker": product.ticker, "Date": date(2026, 1, 5), "Quantité": 10,
        "Prix unitaire": 45.2, "Devise": "EUR", "Taux de change": 1.0,
        "Courtage (EUR)": 0, "TTF (EUR)": 0,
    }
    base.update(overrides)
    return base


@pytest.mark.asyncio
async def test_validate_missing_transactions_sheet_returns_400(client, db_session):
    content = _make_xlsx([], sheet_name="Autre")
    r = await client.post("/api/transactions/import/validate", files=_upload_files(content))
    assert r.status_code == 400
    assert "Transactions" in r.json()["detail"]


@pytest.mark.asyncio
async def test_validate_corrupt_non_xlsx_file_returns_400_not_500(client, db_session):
    """A non-.xlsx upload (wrong file type, corrupted file) must fail cleanly with a 400 —
    openpyxl raises zipfile.BadZipFile/InvalidFileException for this, neither a ValueError
    subclass, so this exercises the normalization in parse_uploaded_workbook."""
    r = await client.post("/api/transactions/import/validate", files=_upload_files(b"not an excel file at all"))
    assert r.status_code == 400
    assert "illisible" in r.json()["detail"]


@pytest.mark.asyncio
async def test_validate_empty_file_returns_zero_rows(client, db_session):
    content = _make_xlsx([])
    r = await client.post("/api/transactions/import/validate", files=_upload_files(content))
    assert r.status_code == 200
    body = r.json()
    assert body["summary"]["total_rows"] == 0
    assert body["rows"] == []


@pytest.mark.asyncio
async def test_validate_skips_fully_blank_rows(client, db_session):
    suffix = id(db_session)
    portfolio, broker, product = await _seed_basic(db_session, suffix)
    wb_rows = [_row(portfolio, broker, product)]
    content = _make_xlsx(wb_rows)
    # Manually inject a blank row between header and data to confirm it's skipped
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content))
    ws = wb["Transactions"]
    ws.insert_rows(2)  # blank row right after header
    buf = io.BytesIO()
    wb.save(buf)

    r = await client.post("/api/transactions/import/validate", files=_upload_files(buf.getvalue()))
    assert r.status_code == 200
    body = r.json()
    assert body["summary"]["total_rows"] == 1
    assert body["rows"][0]["status"] == "ok"


@pytest.mark.asyncio
async def test_validate_zero_byte_file_returns_400_not_500(client, db_session):
    r = await client.post("/api/transactions/import/validate", files=_upload_files(b""))
    assert r.status_code == 400
    assert "illisible" in r.json()["detail"]


@pytest.mark.asyncio
async def test_validate_header_with_stray_whitespace_still_resolves(client, db_session):
    """A header cell like "Portefeuille " (trailing space — e.g. a manually retyped header)
    must not make every row in that column silently read as missing. Found via live QA
    testing alongside the equivalent data-cell whitespace fix."""
    suffix = id(db_session)
    portfolio, broker, product = await _seed_basic(db_session, suffix)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    padded_columns = [f" {col} " for col in TRANSACTION_COLUMNS]
    ws.append(padded_columns)
    ws.append([_row(portfolio, broker, product).get(col) for col in TRANSACTION_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)

    r = await client.post("/api/transactions/import/validate", files=_upload_files(buf.getvalue()))
    assert r.status_code == 200
    body = r.json()
    assert body["rows"][0]["status"] == "ok", body["rows"][0]["errors"]


@pytest.mark.asyncio
async def test_validate_row_with_fewer_cells_than_header_reports_missing_fields(client, db_session):
    """A truncated/ragged row (fewer cells than the header row — e.g. trailing columns left
    empty and stripped by Excel) must degrade to clean "missing field" errors, never crash."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(TRANSACTION_COLUMNS)
    ws.append(["Portfolio1", "Degiro", "Achat", "STN.PA"])  # only 4 of 11 columns
    buf = io.BytesIO()
    wb.save(buf)

    r = await client.post("/api/transactions/import/validate", files=_upload_files(buf.getvalue()))
    assert r.status_code == 200
    body = r.json()
    assert body["rows"][0]["status"] == "error"
    assert any("Date manquante" in e for e in body["rows"][0]["errors"])


@pytest.mark.asyncio
async def test_validate_unevaluated_formula_cell_reports_missing_field(client, db_session):
    """A formula cell with no cached value (e.g. a file built programmatically, or opened but
    never recalculated) reads as None via data_only=True — must degrade to a clean validation
    error on that field, not crash."""
    suffix = id(db_session)
    portfolio, broker, product = await _seed_basic(db_session, suffix)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(TRANSACTION_COLUMNS)
    row = _row(portfolio, broker, product)
    values = [row.get(col) for col in TRANSACTION_COLUMNS]
    values[TRANSACTION_COLUMNS.index("Quantité")] = "=5*2"
    ws.append(values)
    buf = io.BytesIO()
    wb.save(buf)

    r = await client.post("/api/transactions/import/validate", files=_upload_files(buf.getvalue()))
    assert r.status_code == 200
    body = r.json()
    assert body["rows"][0]["status"] == "error"
    assert any("Quantité" in e for e in body["rows"][0]["errors"])


@pytest.mark.asyncio
async def test_validate_detects_in_file_duplicate(client, db_session):
    suffix = id(db_session)
    portfolio, broker, product = await _seed_basic(db_session, suffix)
    content = _make_xlsx([
        _row(portfolio, broker, product),
        _row(portfolio, broker, product),  # exact duplicate of row above
    ])
    r = await client.post("/api/transactions/import/validate", files=_upload_files(content))
    assert r.status_code == 200
    body = r.json()
    assert body["summary"]["ok"] == 1
    assert body["summary"]["duplicates"] == 1
    dup_row = next(row for row in body["rows"] if row["status"] == "duplicate")
    assert dup_row["duplicate_of"]["kind"] == "file"
    assert dup_row["duplicate_of"]["row_number"] == 2


@pytest.mark.asyncio
async def test_validate_detects_in_db_duplicate(client, db_session):
    suffix = id(db_session)
    portfolio, broker, product = await _seed_basic(db_session, suffix)
    existing = Transaction(
        portfolio_id=portfolio.id, account_id=broker.id, date=date(2026, 1, 5),
        type="Actif", operation="Achat", ticker=product.ticker, currency="EUR",
        exchange_rate=1.0, quantity=-10.0, unit_price=45.2,
        unit_price_eur=45.2, total_amount=-452.0, total_amount_eur=-452.0,
    )
    db_session.add(existing)
    await db_session.flush()

    content = _make_xlsx([_row(portfolio, broker, product)])
    r = await client.post("/api/transactions/import/validate", files=_upload_files(content))
    assert r.status_code == 200
    body = r.json()
    assert body["summary"]["duplicates"] == 1
    dup_row = body["rows"][0]
    assert dup_row["status"] == "duplicate"
    assert dup_row["duplicate_of"]["kind"] == "db"
    assert dup_row["duplicate_of"]["transaction_id"] == existing.id


@pytest.mark.asyncio
async def test_validate_does_not_flag_different_amount_as_duplicate(client, db_session):
    suffix = id(db_session)
    portfolio, broker, product = await _seed_basic(db_session, suffix)
    existing = Transaction(
        portfolio_id=portfolio.id, account_id=broker.id, date=date(2026, 1, 5),
        type="Actif", operation="Achat", ticker=product.ticker, currency="EUR",
        exchange_rate=1.0, quantity=-99.0, unit_price=45.2,
        unit_price_eur=45.2, total_amount=-4474.8, total_amount_eur=-4474.8,
    )
    db_session.add(existing)
    await db_session.flush()

    content = _make_xlsx([_row(portfolio, broker, product)])  # quantity 10, not 99
    r = await client.post("/api/transactions/import/validate", files=_upload_files(content))
    body = r.json()
    assert body["rows"][0]["status"] == "ok"


@pytest.mark.asyncio
async def test_validate_row_with_error_is_not_checked_for_duplicates(client, db_session):
    suffix = id(db_session)
    portfolio, broker, product = await _seed_basic(db_session, suffix)
    content = _make_xlsx([_row(portfolio, broker, product, Portefeuille="Ghost")])
    r = await client.post("/api/transactions/import/validate", files=_upload_files(content))
    body = r.json()
    assert body["rows"][0]["status"] == "error"
    assert body["rows"][0]["duplicate_of"] is None


@pytest.mark.asyncio
async def test_validate_never_writes_to_db(client, db_session):
    """A pure dry-run: even for otherwise-valid rows, /validate must never insert anything."""
    suffix = id(db_session)
    portfolio, broker, product = await _seed_basic(db_session, suffix)
    content = _make_xlsx([_row(portfolio, broker, product)])
    r = await client.post("/api/transactions/import/validate", files=_upload_files(content))
    assert r.status_code == 200
    assert r.json()["rows"][0]["status"] == "ok"

    from sqlalchemy import select
    count = (await db_session.execute(
        select(Transaction).where(Transaction.portfolio_id == portfolio.id)
    )).scalars().all()
    assert count == []
