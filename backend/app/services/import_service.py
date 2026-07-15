"""
Bulk transaction import from an uploaded Excel file.

Two-phase design, mirroring the API surface (transaction_import.py router):
  - `resolve_row` is a pure function (no DB access) that maps one raw spreadsheet row to a
    fully-resolved, signed set of Transaction fields or a list of validation errors. It is the
    highest-risk piece of this feature — every "Sens" value's mapping to type/operation/sign is
    spelled out explicitly in SENS_RULES, never inferred at call time.
  - `validate_import` does the DB-touching read work: loading reference data
    (portfolios/brokers/products) once per call, resolving every row against it, then running
    duplicate detection (in-file and in-DB). It never writes to the database.
  - `parse_uploaded_workbook`/`build_template_workbook` handle the .xlsx <-> row-dict
    conversion for both directions (upload parsing, template generation).

Committing the validated rows (calling `create_transaction_core` row by row inside a single
atomic DB transaction) lives in the router (transaction_import.py), not here — it needs
`create_transaction_core`/`_trigger_snapshot_recompute` from app.api.routers.transactions,
and importing a router module here would invert this module's dependency direction.

See CLAUDE.md's "Bulk transaction import (Excel)" section for the full Sens table and the
non-obvious conventions it encodes (Or physique quantity=±1, forex-ticker Devise special case).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date as Date, datetime
from typing import Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Broker, Portfolio, PortfolioAccount, Product, Transaction

FOREX_TICKER_RE = re.compile(r"^([A-Z]{3})[A-Z]{3}=X$")
FLOAT_TOL = 0.005

ASSET_INSTRUMENT_TYPES = {"ETF", "SICAV/FCP", "Action", "Obligation"}
OR_PHYSIQUE_INSTRUMENT_TYPE = "Or physique"
CASH_INSTRUMENT_TYPE = "Cash"


@dataclass(frozen=True)
class SensRule:
    type: str
    operation: Optional[str]
    qty_sign: int                       # +1 or -1, applied to abs(Quantité)
    qty_fixed: Optional[float]          # e.g. 1 for Or physique/Frais — Quantité column must equal this (or be blank)
    price_mode: str                     # "as_entered" | "forced_one" | "optional_zero"
    courtage_allowed: bool
    ttf_allowed: bool
    ticker_kind: str                    # "asset" | "or_physique" | "cash" | "frais"


SENS_RULES: dict[str, SensRule] = {
    "Achat": SensRule("Actif", "Achat", -1, None, "as_entered", True, True, "asset"),
    "Vente": SensRule("Actif", "Vente", 1, None, "as_entered", True, False, "asset"),
    "Achat Or physique": SensRule("Actif", "Achat", -1, 1, "as_entered", True, True, "or_physique"),
    "Vente Or physique": SensRule("Actif", "Vente", 1, 1, "as_entered", True, False, "or_physique"),
    "Attribution": SensRule("Actif", "Attribution", -1, None, "optional_zero", False, False, "asset"),
    "Dépôt": SensRule("Actif", None, 1, None, "forced_one", False, False, "cash"),
    "Retrait": SensRule("Actif", None, -1, None, "forced_one", True, False, "cash"),
    "Revenu": SensRule("Revenu", None, 1, None, "as_entered", False, False, "asset"),
    "Frais": SensRule("Frais", None, -1, 1, "as_entered", False, False, "frais"),
}

TICKER_KIND_LABELS = {
    "asset": "ETF / SICAV/FCP / Action / Obligation",
    "or_physique": "Or physique",
    "cash": "Cash",
    "frais": "Frais",
}


@dataclass
class ResolvedTransaction:
    portfolio_id: int
    account_id: int
    portfolio_name: str
    account_name: str
    date: Date
    type: str
    operation: Optional[str]
    ticker: str
    currency: str
    exchange_rate: float
    quantity: float
    unit_price: float
    courtage_eur: float
    ttf_eur: float


@dataclass
class DuplicateRef:
    kind: str            # "db" | "file"
    transaction_id: Optional[int] = None
    row_number: Optional[int] = None


@dataclass
class RowResult:
    row_number: int
    status: str           # "ok" | "error" | "duplicate"
    sens: Optional[str]
    resolved: Optional[ResolvedTransaction] = None
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    duplicate_of: Optional[DuplicateRef] = None


@dataclass
class ImportSummary:
    total_rows: int
    ok: int
    errors: int
    duplicates: int


@dataclass
class ValidateResult:
    rows: list[RowResult]
    summary: ImportSummary


@dataclass
class ReferenceData:
    portfolios_by_name: dict[str, Portfolio]
    brokers_by_name: dict[str, list[Broker]]
    portfolio_account_pairs: set[tuple[int, int]]   # (portfolio_id, broker_id)
    products_by_ticker: dict[str, Product]


async def load_reference_data(db: AsyncSession) -> ReferenceData:
    portfolios = (await db.execute(select(Portfolio))).scalars().all()
    brokers = (await db.execute(select(Broker))).scalars().all()
    portfolio_accounts = (await db.execute(select(PortfolioAccount))).scalars().all()
    products = (await db.execute(select(Product))).scalars().all()

    brokers_by_name: dict[str, list[Broker]] = {}
    for b in brokers:
        brokers_by_name.setdefault(b.name, []).append(b)

    return ReferenceData(
        portfolios_by_name={p.name: p for p in portfolios},
        brokers_by_name=brokers_by_name,
        portfolio_account_pairs={(pa.portfolio_id, pa.broker_id) for pa in portfolio_accounts},
        products_by_ticker={p.ticker: p for p in products},
    )


def _parse_float(value: object) -> Optional[float]:
    """Returns None for a blank cell, a float otherwise, or raises ValueError."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return float(str(value).strip().replace(",", "."))


def _parse_date(value: object) -> Optional[Date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, Date):
        return value
    return Date.fromisoformat(str(value).strip())


def _expected_currency(ticker: str, product: Product) -> str:
    """A forex-position ticker (JPYEUR=X, USDEUR=X...) is held in its own 3-letter prefix
    currency, not Product.currency (stored as EUR, the reference currency) — see
    TransactionsPage.tsx's handleTickerChange and CLAUDE.md's forex-position section."""
    match = FOREX_TICKER_RE.match(ticker)
    return match.group(1) if match else product.currency


def _display(value: object) -> str:
    """Renders a raw cell value for an error message — a blank cell reads as Python None,
    which must never surface as the literal text "None" in a user-facing error."""
    return "(vide)" if value is None or value == "" else str(value)


def _clean_str(value: object) -> object:
    """Strips leading/trailing whitespace from a text cell — copy-pasting a broker statement
    into Excel commonly introduces stray spaces, which would otherwise fail an exact-match
    lookup (Portefeuille/Compte/Ticker/Devise) even though the entity genuinely exists.
    Non-string values (numbers, dates, None) pass through unchanged."""
    return value.strip() if isinstance(value, str) else value


def _parse_numeric_field(raw: dict, field: str, errors: list[str]) -> Optional[float]:
    """Parses one numeric cell, appending a user-facing error (and returning None) on failure
    instead of raising — shared by every numeric column (Quantité, Prix unitaire, Taux de
    change, Courtage, TTF) so the same accept/reject wording and comma-decimal handling
    (_parse_float) applies uniformly across all five."""
    try:
        return _parse_float(raw.get(field))
    except ValueError:
        errors.append(f"{field} '{raw.get(field)}' invalide.")
        return None


def resolve_row(row_number: int, raw: dict, ref: ReferenceData) -> RowResult:
    """Pure function: maps one raw spreadsheet row to a resolved transaction or errors.
    Never touches the database — duplicate detection happens separately, after every row
    has been independently resolved."""
    errors: list[str] = []
    warnings: list[str] = []

    sens = _clean_str(raw.get("Sens"))
    rule = SENS_RULES.get(sens) if isinstance(sens, str) else None
    if rule is None:
        errors.append(
            f"Sens '{_display(sens)}' invalide — valeurs autorisées : {', '.join(SENS_RULES)}."
        )

    portfolio_name = _clean_str(raw.get("Portefeuille"))
    portfolio = ref.portfolios_by_name.get(portfolio_name) if portfolio_name else None
    if portfolio is None:
        errors.append(
            f"Portefeuille '{_display(portfolio_name)}' introuvable — créez-le d'abord dans l'application."
        )

    broker_name = _clean_str(raw.get("Compte"))
    brokers = ref.brokers_by_name.get(broker_name, []) if broker_name else []
    broker: Optional[Broker] = None
    if not brokers:
        errors.append(
            f"Compte '{_display(broker_name)}' introuvable — créez-le d'abord dans l'application."
        )
    elif len(brokers) > 1:
        errors.append(f"Compte '{_display(broker_name)}' ambigu — plusieurs comptes portent ce nom.")
    else:
        broker = brokers[0]
        if portfolio is not None and (portfolio.id, broker.id) not in ref.portfolio_account_pairs:
            errors.append(
                f"Le compte '{_display(broker_name)}' n'est pas rattaché au portefeuille '{_display(portfolio_name)}'."
            )
            broker = None

    ticker = _clean_str(raw.get("Ticker"))
    product = ref.products_by_ticker.get(ticker) if ticker else None
    if product is None:
        errors.append(
            f"Ticker '{_display(ticker)}' introuvable — créez-le d'abord dans Produits et frais."
        )

    if rule is not None and product is not None:
        kind = rule.ticker_kind
        compatible = (
            (kind == "asset" and product.category == "Actif" and product.instrument_type in ASSET_INSTRUMENT_TYPES)
            or (kind == "or_physique" and product.category == "Actif" and product.instrument_type == OR_PHYSIQUE_INSTRUMENT_TYPE)
            or (kind == "cash" and product.category == "Actif" and product.instrument_type == CASH_INSTRUMENT_TYPE)
            or (kind == "frais" and product.category == "Frais")
        )
        if not compatible:
            errors.append(
                f"Le Sens '{_display(sens)}' n'est pas compatible avec le ticker '{_display(ticker)}' "
                f"(attendu : {TICKER_KIND_LABELS[kind]})."
            )

    quantity_input = _parse_numeric_field(raw, "Quantité", errors)
    price_input = _parse_numeric_field(raw, "Prix unitaire", errors)
    rate_input = _parse_numeric_field(raw, "Taux de change", errors)
    courtage_input = _parse_numeric_field(raw, "Courtage (EUR)", errors) or 0.0
    ttf_input = _parse_numeric_field(raw, "TTF (EUR)", errors) or 0.0

    try:
        tx_date = _parse_date(raw.get("Date"))
        if tx_date is None:
            errors.append("Date manquante.")
    except ValueError:
        errors.append(f"Date '{raw.get('Date')}' invalide.")
        tx_date = None

    devise = _clean_str(raw.get("Devise"))

    if rule is not None:
        if rule.qty_fixed is not None:
            if quantity_input is not None and abs(quantity_input - rule.qty_fixed) > FLOAT_TOL:
                errors.append(f"Quantité doit valoir {rule.qty_fixed} (ou être vide) pour le Sens '{sens}'.")
            quantity_input = rule.qty_fixed
        elif quantity_input is None or quantity_input <= 0:
            errors.append("Quantité doit être un nombre strictement positif.")

        if rule.price_mode == "forced_one":
            if price_input is not None and abs(price_input - 1.0) > FLOAT_TOL:
                errors.append(f"Prix unitaire doit valoir 1.0 (ou être vide) pour le Sens '{sens}'.")
            price_input = 1.0
        elif rule.price_mode == "optional_zero":
            if price_input is None:
                price_input = 0.0
            elif price_input < 0:
                errors.append("Prix unitaire ne peut pas être négatif.")
        else:  # as_entered
            if price_input is None or price_input <= 0:
                errors.append("Prix unitaire doit être un nombre strictement positif.")

        if courtage_input < 0:
            errors.append("Courtage (EUR) ne peut pas être négatif.")
        elif courtage_input > 0 and not rule.courtage_allowed:
            errors.append(f"Courtage (EUR) doit être nul pour le Sens '{sens}'.")

        if ttf_input < 0:
            errors.append("TTF (EUR) ne peut pas être négatif.")
        elif ttf_input > 0 and not rule.ttf_allowed:
            errors.append(f"TTF (EUR) doit être nul pour le Sens '{sens}'.")
        elif ttf_input > 0 and product is not None and not product.is_ttf_eligible:
            warnings.append(
                f"TTF (EUR) renseignée mais le produit '{ticker}' n'est pas marqué éligible TTF."
            )

    if product is not None:
        expected_currency = _expected_currency(ticker, product)
        if devise != expected_currency:
            errors.append(
                f"Devise '{_display(devise)}' incorrecte pour '{ticker}' — attendu '{expected_currency}'."
            )
        if rate_input is None:
            if expected_currency == "EUR":
                rate_input = 1.0
            else:
                errors.append("Taux de change requis pour une devise non-EUR.")
        elif rate_input <= 0:
            errors.append("Taux de change doit être strictement positif.")
        elif expected_currency == "EUR" and abs(rate_input - 1.0) > FLOAT_TOL:
            errors.append("Taux de change doit valoir 1.0 pour une transaction en EUR.")

    if errors:
        return RowResult(row_number=row_number, status="error", sens=sens, errors=errors, warnings=warnings)

    resolved = ResolvedTransaction(
        portfolio_id=portfolio.id,
        account_id=broker.id,
        portfolio_name=portfolio.name,
        account_name=broker.name,
        date=tx_date,
        type=rule.type,
        operation=rule.operation,
        ticker=ticker,
        currency=devise,
        exchange_rate=rate_input,
        quantity=rule.qty_sign * abs(quantity_input),
        unit_price=price_input,
        courtage_eur=courtage_input,
        ttf_eur=ttf_input,
    )
    return RowResult(row_number=row_number, status="ok", sens=sens, resolved=resolved, warnings=warnings)


def _dedup_key(r: ResolvedTransaction) -> tuple:
    return (
        r.portfolio_id, r.account_id, r.date, r.ticker, r.operation,
        round(r.quantity, 6), round(r.unit_price, 6), r.currency,
    )


def _keys_match(a: tuple, b: tuple) -> bool:
    if a[:5] != b[:5] or a[7] != b[7]:
        return False
    return abs(a[5] - b[5]) < FLOAT_TOL and abs(a[6] - b[6]) < FLOAT_TOL


async def _find_db_duplicate(db: AsyncSession, key: tuple) -> Optional[int]:
    portfolio_id, account_id, tx_date, ticker, operation, quantity, unit_price, currency = key
    stmt = select(Transaction).where(
        Transaction.portfolio_id == portfolio_id,
        Transaction.account_id == account_id,
        Transaction.date == tx_date,
        Transaction.ticker == ticker,
        Transaction.currency == currency,
    )
    if operation is None:
        stmt = stmt.where(Transaction.operation.is_(None))
    else:
        stmt = stmt.where(Transaction.operation == operation)
    candidates = (await db.execute(stmt)).scalars().all()
    for tx in candidates:
        if abs(tx.quantity - quantity) < FLOAT_TOL and abs(tx.unit_price - unit_price) < FLOAT_TOL:
            return tx.id
    return None


async def validate_import(db: AsyncSession, raw_rows: list[dict]) -> ValidateResult:
    ref = await load_reference_data(db)
    results = [resolve_row(i, row, ref) for i, row in enumerate(raw_rows, start=2)]

    seen_keys: dict[tuple, int] = {}
    for r in results:
        if r.status != "ok":
            continue
        key = _dedup_key(r.resolved)
        matched_key = next((k for k in seen_keys if _keys_match(k, key)), None)
        if matched_key is not None:
            r.status = "duplicate"
            r.duplicate_of = DuplicateRef(kind="file", row_number=seen_keys[matched_key])
        else:
            seen_keys[key] = r.row_number

    for r in results:
        if r.status != "ok":
            continue
        db_match = await _find_db_duplicate(db, _dedup_key(r.resolved))
        if db_match is not None:
            r.status = "duplicate"
            r.duplicate_of = DuplicateRef(kind="db", transaction_id=db_match)

    summary = ImportSummary(
        total_rows=len(results),
        ok=sum(1 for r in results if r.status == "ok"),
        errors=sum(1 for r in results if r.status == "error"),
        duplicates=sum(1 for r in results if r.status == "duplicate"),
    )
    return ValidateResult(rows=results, summary=summary)


TRANSACTION_COLUMNS = [
    "Portefeuille", "Compte", "Sens", "Ticker", "Date", "Quantité",
    "Prix unitaire", "Devise", "Taux de change", "Courtage (EUR)", "TTF (EUR)",
]


def parse_uploaded_workbook(file_bytes: bytes) -> list[dict]:
    """Parses the uploaded .xlsx into a list of raw row dicts keyed by the "Transactions"
    sheet's own header row (not TRANSACTION_COLUMNS) — a renamed header degrades to "value
    missing" errors per row rather than crashing, since resolve_row only ever does raw.get(...).

    Raises ValueError (never a raw openpyxl/zipfile exception) for any file the router should
    report as a clean 400 — a non-.xlsx upload (wrong file type, corrupted file, plain text)
    fails inside openpyxl with zipfile.BadZipFile or InvalidFileException, neither of which is
    a ValueError subclass; both are normalized here into one user-facing error message.
    """
    from io import BytesIO
    from zipfile import BadZipFile

    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except (BadZipFile, InvalidFileException, KeyError):
        raise ValueError("Fichier illisible — un fichier .xlsx valide est requis.")
    if "Transactions" not in wb.sheetnames:
        raise ValueError("Feuille 'Transactions' introuvable dans le fichier.")
    ws = wb["Transactions"]
    rows = list(ws.iter_rows(min_row=1))
    if not rows:  # pragma: no cover — dead code: openpyxl always reports at least 1 row, even for a blank sheet
        return []
    # Stripped so a header cell with a stray trailing/leading space (e.g. "Portefeuille ",
    # from a manually retyped or re-pasted header) still matches raw.get("Portefeuille")
    # exactly — otherwise every row in that column silently reads as missing.
    header = [_clean_str(cell.value) for cell in rows[0]]
    raw_rows = []
    for row in rows[1:]:
        values = [cell.value for cell in row]
        if all(v is None for v in values):
            continue
        raw_rows.append(dict(zip(header, values)))
    return raw_rows


def _pick_product(ref: ReferenceData, predicate) -> Optional[Product]:
    return next((p for p in ref.products_by_ticker.values() if predicate(p)), None)


def _pick_account(ref: ReferenceData) -> tuple[Optional[Portfolio], Optional[Broker]]:
    portfolios_by_id = {p.id: p for p in ref.portfolios_by_name.values()}
    brokers_by_id = {b.id: b for brokers in ref.brokers_by_name.values() for b in brokers}
    for portfolio_id, broker_id in sorted(ref.portfolio_account_pairs):
        return portfolios_by_id[portfolio_id], brokers_by_id[broker_id]
    return None, None


def _build_example_rows(ref: ReferenceData) -> list[dict]:
    """Best-effort example rows built from live data — never assumes every instrument_type
    exists (a fresh install may have few/no products yet), so each block is independently
    skipped when no matching product/account is found rather than raising. This also makes
    the template a genuine self-consistency smoke test: every row it produces must pass
    resolve_row/validate_import with status="ok" (see test_import_service.py)."""
    portfolio, broker = _pick_account(ref)
    if portfolio is None or broker is None:
        return []

    today = Date.today()
    rows: list[dict] = []

    def add(sens: str, product: Product, quantity: float, price: float,
            currency: Optional[str] = None, rate: float = 1.0,
            courtage: float = 0.0, ttf: float = 0.0) -> None:
        rows.append({
            "Portefeuille": portfolio.name, "Compte": broker.name, "Sens": sens,
            "Ticker": product.ticker, "Date": today,
            "Quantité": quantity, "Prix unitaire": price,
            "Devise": currency or product.currency, "Taux de change": rate,
            "Courtage (EUR)": courtage, "TTF (EUR)": ttf,
        })

    eur_asset = _pick_product(
        ref, lambda p: p.category == "Actif" and p.instrument_type in ASSET_INSTRUMENT_TYPES and p.currency == "EUR"
    )
    if eur_asset:
        add("Achat", eur_asset, 10, 45.20, courtage=2.5)
        add("Vente", eur_asset, 5, 47.10, courtage=2.5)
        add("Attribution", eur_asset, 3, 0.0)
        add("Revenu", eur_asset, 12, 1.35)

    non_eur_asset = _pick_product(
        ref, lambda p: p.category == "Actif" and p.instrument_type in ASSET_INSTRUMENT_TYPES and p.currency != "EUR"
    )
    if non_eur_asset:
        add("Achat", non_eur_asset, 8, 120.0, currency=non_eur_asset.currency, rate=0.92, courtage=3.0)

    cash_eur = _pick_product(
        ref, lambda p: p.category == "Actif" and p.instrument_type == CASH_INSTRUMENT_TYPE
        and not FOREX_TICKER_RE.match(p.ticker)
    )
    if cash_eur:
        add("Dépôt", cash_eur, 1000, 1.0, currency="EUR")
        add("Retrait", cash_eur, 200, 1.0, currency="EUR", courtage=5.0)

    forex_cash = _pick_product(
        ref, lambda p: p.category == "Actif" and p.instrument_type == CASH_INSTRUMENT_TYPE
        and FOREX_TICKER_RE.match(p.ticker)
    )
    if forex_cash:
        forex_currency = FOREX_TICKER_RE.match(forex_cash.ticker).group(1)
        add("Dépôt", forex_cash, 50000, 1.0, currency=forex_currency, rate=0.0061)

    or_physique = _pick_product(
        ref, lambda p: p.category == "Actif" and p.instrument_type == OR_PHYSIQUE_INSTRUMENT_TYPE
    )
    if or_physique:
        add("Achat Or physique", or_physique, 1, 1850.0)
        add("Vente Or physique", or_physique, 1, 2100.0)

    frais_product = _pick_product(ref, lambda p: p.category == "Frais")
    if frais_product:
        add("Frais", frais_product, 1, 12.0)

    return rows


_INSTRUCTIONS_LINES = [
    ("Import de transactions — instructions", True),
    ("", False),
    ("Créez d'abord tout portefeuille, compte (broker) ou produit manquant via l'application "
     "avant l'import — une ligne référençant un élément inconnu sera rejetée.", False),
    ("", False),
    ("Colonne 'Sens' — détermine automatiquement le type, l'opération et le signe interne :", True),
    ("Achat — achat d'un ETF/SICAV-FCP/Action/Obligation. Quantité et Prix unitaire positifs.", False),
    ("Vente — vente du même type de produit.", False),
    ("Achat Or physique / Vente Or physique — Quantité toujours 1 (ou vide) ; 'Prix unitaire' "
     "représente la VALEUR TOTALE payée/reçue, pas un prix au gramme/once. Exemple : achat d'un "
     "lingot pour 1 850,00 € -> Quantité = 1, Prix unitaire = 1850.", False),
    ("Attribution — don d'actions gratuit. Prix unitaire optionnel (valeur de marché ou 0). "
     "Courtage et TTF doivent être nuls.", False),
    ("Dépôt / Retrait — mouvement de cash (compte espèces ou position devise). Prix unitaire "
     "toujours 1.0. Un Retrait peut porter un Courtage (frais de retrait broker) ; jamais de TTF.", False),
    ("Revenu — dividende ou revenu perçu sur un actif détenu.", False),
    ("Frais — frais isolé non lié à un achat/vente (ex. tenue de compte). Quantité toujours 1 "
     "(ou vide), Prix unitaire = montant du frais.", False),
    ("", False),
    ("Devise — cas particulier des positions de change (ex. JPYEUR=X) : la devise attendue est "
     "le préfixe à 3 lettres du ticker (JPY), pas la devise de référence du produit (EUR).", False),
    ("Taux de change — obligatoire et strictement positif pour toute devise non-EUR ; doit "
     "valoir 1.0 en EUR.", False),
    ("Courtage (EUR) / TTF (EUR) — recopiez les montants réels de votre relevé broker ; jamais "
     "recalculés automatiquement par l'application.", False),
    ("", False),
    ("Le fichier n'a pas besoin d'être trié par date — l'application réordonne automatiquement "
     "avant l'import.", False),
    ("Les lignes déjà présentes en base (même portefeuille/compte/ticker/sens/date/quantité/"
     "prix/devise) sont signalées comme doublons ; vous choisissez de les importer quand même "
     "ou non au moment de la confirmation.", False),
]


def _fill_instructions_sheet(ws) -> None:
    from openpyxl.styles import Font

    bold = Font(bold=True)
    for row_idx, (text, is_bold) in enumerate(_INSTRUCTIONS_LINES, start=1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        if is_bold:
            cell.font = bold
    ws.column_dimensions["A"].width = 110


def build_template_workbook(ref: ReferenceData):
    """Builds the downloadable .xlsx template in memory: a "Transactions" sheet (headers +
    live-data example rows + dropdowns) and a separate "Instructions" sheet. Never leaves the
    user's machine (personal, local single-user app) so using real portfolio/broker/ticker
    names in the examples is safe and makes them genuinely importable, unlike fake placeholders.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(TRANSACTION_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for idx, col in enumerate(TRANSACTION_COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = max(14, len(col) + 2)

    for row in _build_example_rows(ref):
        ws.append([row[col] for col in TRANSACTION_COLUMNS])

    instructions_ws = wb.create_sheet("Instructions")
    _fill_instructions_sheet(instructions_ws)

    lists_ws = wb.create_sheet("_Lists")
    lists_ws.sheet_state = "hidden"
    portfolio_names = sorted(ref.portfolios_by_name.keys())
    broker_names = sorted(ref.brokers_by_name.keys())
    ticker_names = sorted(ref.products_by_ticker.keys())
    for col_idx, values in enumerate((portfolio_names, broker_names, ticker_names), start=1):
        for row_idx, value in enumerate(values, start=1):
            lists_ws.cell(row=row_idx, column=col_idx, value=value)

    max_rows = 500

    def add_dropdown(col_letter: str, formula: str) -> None:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{max_rows}")

    def lists_ref(lists_col_letter: str, count: int) -> str:
        return f"='_Lists'!${lists_col_letter}$1:${lists_col_letter}${max(count, 1)}"

    add_dropdown("A", lists_ref("A", len(portfolio_names)))
    add_dropdown("B", lists_ref("B", len(broker_names)))
    add_dropdown("C", '"{}"'.format(",".join(SENS_RULES.keys())))
    add_dropdown("D", lists_ref("C", len(ticker_names)))

    return wb
